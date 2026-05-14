"""
Generates the combined PDF + Excel report for weekly/monthly scheduled delivery.
Called by both Celery tasks and APScheduler fallback.
"""
import io
from datetime import date

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, HRFlowable,
    Table, TableStyle, KeepTogether,
)
from reportlab.lib.enums import TA_CENTER, TA_RIGHT

from app.utils.analytics import (
    booking_stats, revenue_stats, active_trucks,
    new_customers, top_routes, booking_trend, customer_growth,
)

ORANGE     = colors.HexColor("#E87820")
LIGHT_GREY = colors.HexColor("#F5F5F5")
GREY       = colors.HexColor("#757575")


# ── PDF ───────────────────────────────────────────────────────────────────────

def _tbl(data, col_widths):
    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0),  ORANGE),
        ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0),  8),
        ("FONTSIZE",      (0, 1), (-1, -1), 7),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, LIGHT_GREY]),
        ("GRID",          (0, 0), (-1, -1), 0.3, colors.HexColor("#E0E0E0")),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING",   (0, 0), (-1, -1), 5),
    ]))
    return t


def generate_report_pdf(db, from_date: date, to_date: date, report_type: str) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=1.5*cm,   bottomMargin=1.5*cm,
    )

    styles = getSampleStyleSheet()
    title_s   = ParagraphStyle("t",   parent=styles["Heading1"], fontSize=18, textColor=ORANGE, spaceAfter=2)
    sub_s     = ParagraphStyle("s",   parent=styles["Normal"],   fontSize=9,  textColor=GREY,   spaceAfter=4)
    section_s = ParagraphStyle("sec", parent=styles["Heading2"], fontSize=11, textColor=ORANGE, spaceBefore=12, spaceAfter=4)
    cell_s    = ParagraphStyle("c",   parent=styles["Normal"],   fontSize=8)
    right_s   = ParagraphStyle("r",   parent=styles["Normal"],   fontSize=9,  alignment=TA_RIGHT)

    period_label = f"{from_date.strftime('%d %b %Y')} — {to_date.strftime('%d %b %Y')}"
    story = []

    # ── header ──
    story.append(Paragraph("GOGOTRUK", title_s))
    story.append(Paragraph(f"{report_type.title()} Business Report  |  {period_label}", sub_s))
    story.append(HRFlowable(width="100%", thickness=2, color=ORANGE, spaceAfter=10))

    pw = landscape(A4)[0] - 3*cm

    # ── summary cards ──
    bk  = booking_stats(db, from_date, to_date)
    rev = revenue_stats(db, from_date, to_date)
    at  = active_trucks(db)
    nc  = new_customers(db, from_date, to_date)

    story.append(Paragraph("Summary", section_s))
    summary_data = [
        ["Metric", "Value"],
        ["Total Bookings",    str(bk["total"])],
        ["Confirmed",         str(bk["confirmed"])],
        ["Completed",         str(bk["completed"])],
        ["Cancelled",         str(bk["cancelled"])],
        ["Total Invoiced",    f"₹{rev['total_invoiced']:,.2f}"],
        ["Collected",         f"₹{rev['total_collected']:,.2f}"],
        ["Outstanding",       f"₹{rev['outstanding']:,.2f}"],
        ["Active Trucks",     str(at)],
        ["New Customers",     str(nc)],
    ]
    story.append(_tbl(summary_data, [pw*0.4, pw*0.3]))
    story.append(Spacer(1, 0.4*cm))

    # ── top routes ──
    routes = top_routes(db, from_date, to_date, limit=10)
    if routes:
        story.append(Paragraph("Top Routes", section_s))
        route_data = [["#", "From", "To", "Bookings"]]
        for i, r in enumerate(routes, 1):
            route_data.append([str(i), r["from"][:40], r["to"][:40], str(r["bookings"])])
        story.append(_tbl(route_data, [pw*0.04, pw*0.38, pw*0.38, pw*0.10]))
        story.append(Spacer(1, 0.4*cm))

    # ── daily trend ──
    trend = booking_trend(db, from_date, to_date)
    if trend:
        story.append(Paragraph("Daily Trend", section_s))
        trend_data = [["Date", "Bookings", "Revenue (₹)"]]
        for row in trend:
            trend_data.append([row["date"], str(row["bookings"]), f"{row['revenue']:,.2f}"])
        story.append(_tbl(trend_data, [pw*0.20, pw*0.15, pw*0.20]))
        story.append(Spacer(1, 0.4*cm))

    # ── customer growth (monthly only) ──
    if report_type == "monthly":
        growth = customer_growth(db, months=6)
        if growth:
            story.append(Paragraph("Customer Growth (last 6 months)", section_s))
            growth_data = [["Month", "New Customers"]]
            for row in growth:
                growth_data.append([row["month"], str(row["new_customers"])])
            story.append(_tbl(growth_data, [pw*0.25, pw*0.20]))

    doc.build(story)
    return buffer.getvalue()


# ── Excel ─────────────────────────────────────────────────────────────────────

def generate_report_excel(db, from_date: date, to_date: date, report_type: str) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    HDR_FONT  = Font(bold=True, color="FFFFFF")
    HDR_FILL  = PatternFill("solid", fgColor="E87820")
    ALT_FILL  = PatternFill("solid", fgColor="FFF3E0")

    def _write_sheet(ws, headers, rows, col_widths=None):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font  = HDR_FONT
            cell.fill  = HDR_FILL
            cell.alignment = Alignment(horizontal="center")
        for row_idx, row in enumerate(rows, 2):
            for col_idx, val in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)
                if row_idx % 2 == 0:
                    ws.cell(row=row_idx, column=col_idx).fill = ALT_FILL
        if col_widths:
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[ws.cell(1, i).column_letter].width = w
        else:
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 18

    wb = openpyxl.Workbook()
    period_label = f"{from_date} to {to_date}"

    # Sheet 1: Summary
    ws = wb.active
    ws.title = "Summary"
    bk  = booking_stats(db, from_date, to_date)
    rev = revenue_stats(db, from_date, to_date)
    at  = active_trucks(db)
    nc  = new_customers(db, from_date, to_date)
    _write_sheet(ws, ["Metric", "Value"], [
        ["Report Type",     report_type.title()],
        ["Period",          period_label],
        ["Total Bookings",  bk["total"]],
        ["Confirmed",       bk["confirmed"]],
        ["Completed",       bk["completed"]],
        ["Cancelled",       bk["cancelled"]],
        ["Rejected",        bk["rejected"]],
        ["Total Invoiced",  rev["total_invoiced"]],
        ["Collected",       rev["total_collected"]],
        ["Outstanding",     rev["outstanding"]],
        ["Active Trucks",   at],
        ["New Customers",   nc],
    ], col_widths=[25, 20])

    # Sheet 2: Daily Trend
    ws2 = wb.create_sheet("Daily Trend")
    trend = booking_trend(db, from_date, to_date)
    _write_sheet(ws2, ["Date", "Bookings", "Revenue (₹)"],
                 [[r["date"], r["bookings"], r["revenue"]] for r in trend],
                 col_widths=[15, 12, 18])

    # Sheet 3: Top Routes
    ws3 = wb.create_sheet("Top Routes")
    routes = top_routes(db, from_date, to_date, limit=10)
    _write_sheet(ws3, ["#", "From", "To", "Bookings"],
                 [[i+1, r["from"], r["to"], r["bookings"]] for i, r in enumerate(routes)],
                 col_widths=[5, 40, 40, 12])

    # Sheet 4: Customer Growth (monthly only)
    if report_type == "monthly":
        ws4 = wb.create_sheet("Customer Growth")
        growth = customer_growth(db, months=6)
        _write_sheet(ws4, ["Month", "New Customers"],
                     [[r["month"], r["new_customers"]] for r in growth],
                     col_widths=[15, 18])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Email delivery ────────────────────────────────────────────────────────────

def email_report(pdf_bytes: bytes, excel_bytes: bytes, report_type: str,
                 from_date: date, to_date: date):
    from app.config import settings

    period = f"{from_date.strftime('%d %b %Y')} – {to_date.strftime('%d %b %Y')}"
    subject = f"GoGoTruk {report_type.title()} Report — {period}"
    recipients = [e.strip() for e in settings.MANAGEMENT_EMAIL.split(",") if e.strip()]

    if not recipients:
        print(f"[report] No MANAGEMENT_EMAIL set — skipping email for {report_type} report")
        return

    if settings.DEV_MODE:
        print(f"\n[report] {report_type.title()} report ready ({period})")
        print(f"[report]   Would email to: {', '.join(recipients)}")
        print(f"[report]   PDF size:   {len(pdf_bytes):,} bytes")
        print(f"[report]   Excel size: {len(excel_bytes):,} bytes\n")
        return

    if not settings.SENDGRID_API_KEY:
        print("[report] SENDGRID_API_KEY not set — cannot send report email")
        return

    try:
        import base64
        from sendgrid import SendGridAPIClient
        from sendgrid.helpers.mail import (
            Mail, Attachment, FileContent, FileName, FileType, Disposition,
        )

        body = (
            f"<p>Hi,</p>"
            f"<p>Please find attached the GoGoTruk <strong>{report_type}</strong> business report "
            f"for the period <strong>{period}</strong>.</p>"
            f"<p>This report includes bookings summary, revenue, top routes, and customer growth.</p>"
            f"<p>— GoGoTruk Automated Reporting</p>"
        )

        message = Mail(
            from_email=settings.SENDGRID_FROM_EMAIL,
            to_emails=recipients,
            subject=subject,
            html_content=body,
        )

        fname = f"gogotruk_{report_type}_report_{from_date}.pdf"
        message.attachment = Attachment(
            FileContent(base64.b64encode(pdf_bytes).decode()),
            FileName(fname),
            FileType("application/pdf"),
            Disposition("attachment"),
        )
        excel_name = f"gogotruk_{report_type}_report_{from_date}.xlsx"
        message.attachment = Attachment(
            FileContent(base64.b64encode(excel_bytes).decode()),
            FileName(excel_name),
            FileType("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            Disposition("attachment"),
        )

        sg = SendGridAPIClient(settings.SENDGRID_API_KEY)
        sg.send(message)
        print(f"[report] {report_type.title()} report emailed to {', '.join(recipients)}")
    except Exception as e:
        print(f"[report] Email send failed: {e}")
