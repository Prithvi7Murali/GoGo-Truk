import io
from datetime import datetime, date
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.kyc import CustomerKYC, OwnerKYC, CompanyKYC
from app.models.notification_log import NotificationLog
from app.models.fleet import Fleet
from app.models.booking import Booking
from app.models.invoice import Invoice
from app.schemas.admin_dashboard import (
    DashboardMetrics, KYCStats, BookingStats, RevenueStats, FleetStats,
    KYCQueueItem, BookingOverviewItem, RevenueReportItem,
)
from app.utils.auth import get_current_admin

router = APIRouter(prefix="/api/admin/dashboard", tags=["Admin Dashboard"])


# ── helpers ─────────────────────────────────────────────────────────────────

def _count(db, model, **filters):
    q = db.query(func.count(model.id))
    for attr, val in filters.items():
        q = q.filter(getattr(model, attr) == val)
    return q.scalar() or 0


# ── metrics ──────────────────────────────────────────────────────────────────

@router.get("/metrics", response_model=DashboardMetrics)
def get_metrics(db: Session = Depends(get_db), _=Depends(get_current_admin)):
    kyc_total     = _count(db, CustomerKYC)
    kyc_pending   = _count(db, CustomerKYC, status="Pending")
    kyc_verified  = _count(db, CustomerKYC, status="Verified")
    kyc_rejected  = _count(db, CustomerKYC, status="Rejected")

    bk_total     = _count(db, Booking)
    bk_pending   = _count(db, Booking, status="Pending")
    bk_confirmed = _count(db, Booking, status="Confirmed")
    bk_completed = _count(db, Booking, status="Completed")
    bk_cancelled = _count(db, Booking, status="Cancelled")
    bk_rejected  = _count(db, Booking, status="Rejected")

    rev_row = db.query(
        func.coalesce(func.sum(Invoice.total_amount), 0.0),
        func.count(Invoice.id),
    ).first()
    paid_row = db.query(
        func.coalesce(func.sum(Invoice.total_amount), 0.0),
    ).filter(Invoice.status == "Sent").first()

    total_invoiced  = float(rev_row[0])
    invoice_count   = int(rev_row[1])
    total_paid      = float(paid_row[0])
    total_outstanding = total_invoiced - total_paid

    fleet_total    = _count(db, Fleet)
    fleet_active   = _count(db, Fleet, is_active=True)
    fleet_inactive = fleet_total - fleet_active

    return DashboardMetrics(
        kyc=KYCStats(
            pending=kyc_pending,
            verified=kyc_verified,
            rejected=kyc_rejected,
            total=kyc_total,
        ),
        bookings=BookingStats(
            pending=bk_pending,
            confirmed=bk_confirmed,
            completed=bk_completed,
            cancelled=bk_cancelled,
            rejected=bk_rejected,
            total=bk_total,
        ),
        revenue=RevenueStats(
            total_invoiced=total_invoiced,
            total_paid=total_paid,
            total_outstanding=total_outstanding,
            invoice_count=invoice_count,
        ),
        fleet=FleetStats(
            total=fleet_total,
            active=fleet_active,
            inactive=fleet_inactive,
        ),
    )


# ── KYC queue ────────────────────────────────────────────────────────────────

@router.get("/kyc-queue", response_model=list[KYCQueueItem])
def kyc_queue(
    status: str | None = Query(None),
    customer_type: str | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    _=Depends(get_current_admin),
):
    q = db.query(CustomerKYC)
    if status:
        q = q.filter(CustomerKYC.status == status)
    if customer_type:
        q = q.filter(CustomerKYC.customer_type == customer_type)
    records = q.order_by(CustomerKYC.created_at.asc()).offset((page - 1) * page_size).limit(page_size).all()
    return [
        KYCQueueItem(
            id=r.id,
            full_name=f"{r.first_name} {r.last_name}",
            customer_type=r.customer_type,
            mobile=r.mobile,
            status=str(r.status.value) if hasattr(r.status, "value") else str(r.status),
            submitted_at=r.created_at,
        )
        for r in records
    ]


# ── Fleet verification queue ─────────────────────────────────────────────────

@router.get("/fleet-queue")
def fleet_queue(
    is_active: bool | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    _=Depends(get_current_admin),
):
    q = db.query(Fleet)
    if is_active is not None:
        q = q.filter(Fleet.is_active == is_active)
    total = q.count()
    records = q.order_by(Fleet.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return {
        "total": total,
        "page": page,
        "results": [
            {
                "id": f.id,
                "registration_number": f.registration_number,
                "vehicle_type": f.vehicle_type,
                "owner_kyc_id": f.owner_kyc_id,
                "is_active": f.is_active,
                "rc_expiry_date": str(f.rc_expiry_date) if f.rc_expiry_date else None,
                "insurance_expiry_date": str(f.insurance_expiry_date) if f.insurance_expiry_date else None,
                "permit_expiry_date": str(f.permit_expiry_date) if f.permit_expiry_date else None,
                "puc_expiry_date": str(f.puc_expiry_date) if f.puc_expiry_date else None,
            }
            for f in records
        ],
    }


# ── Booking overview ─────────────────────────────────────────────────────────

@router.get("/bookings", response_model=list[BookingOverviewItem])
def booking_overview(
    status: str | None = Query(None),
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    _=Depends(get_current_admin),
):
    q = db.query(Booking, CustomerKYC, Fleet).join(
        CustomerKYC, Booking.customer_kyc_id == CustomerKYC.id
    ).join(
        Fleet, Booking.fleet_id == Fleet.id
    ).outerjoin(
        OwnerKYC, Fleet.owner_kyc_id == OwnerKYC.id
    )

    if status:
        q = q.filter(Booking.status == status)
    if date_from:
        q = q.filter(Booking.booking_date >= date_from)
    if date_to:
        q = q.filter(Booking.booking_date <= date_to)

    rows = q.order_by(Booking.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()

    result = []
    for booking, customer, fleet in rows:
        owner = db.query(OwnerKYC).filter(OwnerKYC.id == fleet.owner_kyc_id).first()
        result.append(BookingOverviewItem(
            id=booking.id,
            customer_name=f"{customer.first_name} {customer.last_name}",
            owner_name=f"{owner.first_name} {owner.last_name}" if owner else "—",
            vehicle_number=fleet.registration_number,
            pickup_location=booking.pickup_address,
            drop_location=booking.destination_address,
            pickup_date=str(booking.booking_date),
            status=booking.status,
            created_at=booking.created_at,
        ))
    return result


# ── Revenue report ────────────────────────────────────────────────────────────

@router.get("/revenue", response_model=list[RevenueReportItem])
def revenue_report(
    status: str | None = Query(None),
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    _=Depends(get_current_admin),
):
    q = db.query(Invoice, CustomerKYC).join(
        CustomerKYC, Invoice.customer_kyc_id == CustomerKYC.id
    )
    if status:
        q = q.filter(Invoice.status == status)
    if date_from:
        q = q.filter(func.date(Invoice.created_at) >= date_from)
    if date_to:
        q = q.filter(func.date(Invoice.created_at) <= date_to)

    rows = q.order_by(Invoice.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return [
        RevenueReportItem(
            invoice_id=inv.id,
            invoice_number=inv.invoice_number,
            booking_id=inv.booking_id,
            customer_name=f"{cust.first_name} {cust.last_name}",
            total_amount=inv.total_amount,
            status=inv.status,
            generated_at=inv.created_at,
        )
        for inv, cust in rows
    ]


# ── Excel export ──────────────────────────────────────────────────────────────

@router.get("/export/bookings.xlsx")
def export_bookings_excel(
    status: str | None = Query(None),
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    db: Session = Depends(get_db),
    _=Depends(get_current_admin),
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(status_code=500, detail="openpyxl not installed — run: pip install openpyxl")

    q = db.query(Booking, CustomerKYC, Fleet).join(
        CustomerKYC, Booking.customer_kyc_id == CustomerKYC.id
    ).join(Fleet, Booking.fleet_id == Fleet.id)

    if status:
        q = q.filter(Booking.status == status)
    if date_from:
        q = q.filter(Booking.booking_date >= date_from)
    if date_to:
        q = q.filter(Booking.booking_date <= date_to)

    rows = q.order_by(Booking.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bookings"

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="E87820")
    center_align = Alignment(horizontal="center")

    headers = [
        "Booking ID", "Customer Name", "Vehicle Number", "Pickup Address",
        "Drop Address", "Booking Date", "Goods Type", "Weight (kg)", "Status", "Created At",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for row_idx, (booking, customer, fleet) in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=booking.id)
        ws.cell(row=row_idx, column=2, value=f"{customer.first_name} {customer.last_name}")
        ws.cell(row=row_idx, column=3, value=fleet.registration_number)
        ws.cell(row=row_idx, column=4, value=booking.pickup_address)
        ws.cell(row=row_idx, column=5, value=booking.destination_address)
        ws.cell(row=row_idx, column=6, value=str(booking.booking_date))
        ws.cell(row=row_idx, column=7, value=booking.goods_type)
        ws.cell(row=row_idx, column=8, value=booking.goods_weight_kg)
        ws.cell(row=row_idx, column=9, value=booking.status)
        ws.cell(row=row_idx, column=10, value=str(booking.created_at)[:19] if booking.created_at else "")

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=bookings.xlsx"},
    )


@router.get("/export/revenue.xlsx")
def export_revenue_excel(
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    db: Session = Depends(get_db),
    _=Depends(get_current_admin),
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    q = db.query(Invoice, CustomerKYC).join(CustomerKYC, Invoice.customer_kyc_id == CustomerKYC.id)
    if date_from:
        q = q.filter(func.date(Invoice.created_at) >= date_from)
    if date_to:
        q = q.filter(func.date(Invoice.created_at) <= date_to)

    rows = q.order_by(Invoice.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Revenue"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="E87820")

    headers = [
        "Invoice Number", "Booking ID", "Customer Name",
        "Base Fare", "GST Type", "CGST", "SGST", "IGST",
        "Total Amount", "Status", "Date",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, (inv, cust) in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=inv.invoice_number)
        ws.cell(row=row_idx, column=2, value=inv.booking_id)
        ws.cell(row=row_idx, column=3, value=f"{cust.first_name} {cust.last_name}")
        ws.cell(row=row_idx, column=4, value=inv.base_fare)
        ws.cell(row=row_idx, column=5, value=inv.gst_type)
        ws.cell(row=row_idx, column=6, value=inv.cgst_amount)
        ws.cell(row=row_idx, column=7, value=inv.sgst_amount)
        ws.cell(row=row_idx, column=8, value=inv.igst_amount)
        ws.cell(row=row_idx, column=9, value=inv.total_amount)
        ws.cell(row=row_idx, column=10, value=inv.status)
        ws.cell(row=row_idx, column=11, value=str(inv.created_at)[:10] if inv.created_at else "")

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=revenue.xlsx"},
    )


# ── PDF exports ──────────────────────────────────────────────────────────────

@router.get("/export/kyc.pdf")
def export_kyc_pdf(
    status: str | None = Query(None),
    customer_type: str | None = Query(None),
    db: Session = Depends(get_db),
    _=Depends(get_current_admin),
):
    from app.utils.report_pdf import generate_kyc_pdf
    q = db.query(CustomerKYC)
    if status:
        q = q.filter(CustomerKYC.status == status)
    if customer_type:
        q = q.filter(CustomerKYC.customer_type == customer_type)
    records = q.order_by(CustomerKYC.created_at.asc()).all()
    rows = [
        {
            "full_name": f"{r.first_name} {r.last_name}",
            "customer_type": r.customer_type,
            "mobile": r.mobile,
            "email": r.email,
            "city": r.city,
            "state": r.state,
            "status": str(r.status.value) if hasattr(r.status, "value") else str(r.status),
            "submitted_at": r.created_at,
        }
        for r in records
    ]
    pdf = generate_kyc_pdf(rows)
    return StreamingResponse(
        io.BytesIO(pdf),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=kyc_report.pdf"},
    )


@router.get("/export/bookings.pdf")
def export_bookings_pdf(
    status: str | None = Query(None),
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    db: Session = Depends(get_db),
    _=Depends(get_current_admin),
):
    from app.utils.report_pdf import generate_bookings_pdf
    q = db.query(Booking, CustomerKYC, Fleet).join(
        CustomerKYC, Booking.customer_kyc_id == CustomerKYC.id
    ).join(Fleet, Booking.fleet_id == Fleet.id)
    if status:
        q = q.filter(Booking.status == status)
    if date_from:
        q = q.filter(Booking.booking_date >= date_from)
    if date_to:
        q = q.filter(Booking.booking_date <= date_to)
    rows_db = q.order_by(Booking.created_at.desc()).all()
    rows = []
    for booking, customer, fleet in rows_db:
        owner = db.query(OwnerKYC).filter(OwnerKYC.id == fleet.owner_kyc_id).first()
        rows.append({
            "id": booking.id,
            "customer_name": f"{customer.first_name} {customer.last_name}",
            "owner_name": f"{owner.first_name} {owner.last_name}" if owner else "—",
            "vehicle_number": fleet.registration_number,
            "pickup_address": booking.pickup_address,
            "destination_address": booking.destination_address,
            "booking_date": booking.booking_date,
            "goods_type": booking.goods_type,
            "goods_weight_kg": booking.goods_weight_kg,
            "status": booking.status,
        })
    pdf = generate_bookings_pdf(rows)
    return StreamingResponse(
        io.BytesIO(pdf),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=bookings_report.pdf"},
    )


@router.get("/export/revenue.pdf")
def export_revenue_pdf(
    status: str | None = Query(None),
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    db: Session = Depends(get_db),
    _=Depends(get_current_admin),
):
    from app.utils.report_pdf import generate_revenue_pdf
    q = db.query(Invoice, CustomerKYC).join(CustomerKYC, Invoice.customer_kyc_id == CustomerKYC.id)
    if status:
        q = q.filter(Invoice.status == status)
    if date_from:
        q = q.filter(func.date(Invoice.created_at) >= date_from)
    if date_to:
        q = q.filter(func.date(Invoice.created_at) <= date_to)
    rows_db = q.order_by(Invoice.created_at.desc()).all()

    total_invoiced = sum(inv.total_amount for inv, _ in rows_db)
    total_paid     = sum(inv.total_amount for inv, _ in rows_db if inv.status == "Sent")

    rows = [
        {
            "invoice_number": inv.invoice_number,
            "booking_id": inv.booking_id,
            "customer_name": f"{cust.first_name} {cust.last_name}",
            "base_fare": inv.base_fare,
            "gst_type": inv.gst_type,
            "cgst_amount": inv.cgst_amount,
            "sgst_amount": inv.sgst_amount,
            "igst_amount": inv.igst_amount,
            "total_amount": inv.total_amount,
            "status": inv.status,
            "generated_at": inv.created_at,
        }
        for inv, cust in rows_db
    ]
    totals = {
        "total_invoiced": total_invoiced,
        "total_paid": total_paid,
        "outstanding": total_invoiced - total_paid,
    }
    pdf = generate_revenue_pdf(rows, totals)
    return StreamingResponse(
        io.BytesIO(pdf),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=revenue_report.pdf"},
    )


# ── Notification log ─────────────────────────────────────────────────────────

@router.get("/notifications")
def notification_log(
    event_type: str | None = Query(None),
    channel:    str | None = Query(None),
    status:     str | None = Query(None),
    page:       int = Query(1, ge=1),
    page_size:  int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    _=Depends(get_current_admin),
):
    q = db.query(NotificationLog)
    if event_type:
        q = q.filter(NotificationLog.event_type == event_type)
    if channel:
        q = q.filter(NotificationLog.channel == channel)
    if status:
        q = q.filter(NotificationLog.status == status)
    total = q.count()
    rows  = q.order_by(NotificationLog.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return {
        "total": total,
        "page":  page,
        "results": [
            {
                "id":               r.id,
                "event_type":       r.event_type,
                "channel":          r.channel,
                "recipient_mobile": r.recipient_mobile,
                "recipient_email":  r.recipient_email,
                "subject":          r.subject,
                "message":          r.message,
                "status":           r.status,
                "error_message":    r.error_message,
                "created_at":       r.created_at,
            }
            for r in rows
        ],
    }


# ── User management ──────────────────────────────────────────────────────────

@router.get("/users/kyc/{kyc_id}")
def get_kyc_detail(kyc_id: int, db: Session = Depends(get_db), _=Depends(get_current_admin)):
    record = db.query(CustomerKYC).filter(CustomerKYC.id == kyc_id).first()
    if not record:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="KYC record not found")
    return {
        "id": record.id,
        "full_name": f"{record.first_name} {record.last_name}",
        "email": record.email,
        "mobile": record.mobile,
        "customer_type": record.customer_type,
        "status": str(record.status.value) if hasattr(record.status, "value") else str(record.status),
        "city": record.city,
        "state": record.state,
        "id_proof_url": record.id_proof_url,
        "created_at": record.created_at,
    }
